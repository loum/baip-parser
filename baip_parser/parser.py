# pylint: disable=R0903,C0111,R0902
"""The :class:`baip_parser.Parser` provides file ingest support.

"""
__all__ = ["Parser"]

import os
import openpyxl

from logga.log import log


class Parser(object):
    """:class:`baip_parser.Parser`

    .. attribute:: *filepath*
        fully qualified name of the ``xlsx`` file to parse.

    """
    _filepath = None
    _workbook = None
    _skip_sheets = []
    _cells_to_extract = []
    _parsed_values = {}

    @property
    def filepath(self):
        return self._filepath

    @filepath.setter
    def filepath(self, value):
        self._filepath = value

    @property
    def workbook(self):
        return self._workbook

    @workbook.setter
    def workbook(self, value):
        self._workbook = value

    @property
    def sheet_names(self):
        sheet_names = []

        if self.workbook is not None:
            sheet_names.extend(self.workbook.sheetnames)

        return sheet_names

    @property
    def skip_sheets(self):
        return self._skip_sheets

    @skip_sheets.setter
    def skip_sheets(self, values=None):
        del self._skip_sheets[:]
        self._skip_sheets = []

        if values is not None and isinstance(values, list):
            self._skip_sheets.extend(values)

    @property
    def cells_to_extract(self):
        return self._cells_to_extract

    @cells_to_extract.setter
    def cells_to_extract(self, values=None):
        del self._cells_to_extract[:]
        self._cells_to_extract = []

        if values is not None and isinstance(values, list):
            self._cells_to_extract.extend(values)

    def __init__(self):
        pass

    def open(self, filepath=None):
        """Attempt to open the ``xlsx`` file for processing.

        **Args:**
            *filepath*: override the :attr:`parser.filepath` attribute

        """
        file_to_open = None

        if filepath is not None:
            file_to_open = filepath
        else:
            file_to_open = self.filepath

        if file_to_open is not None:
            log.debug('Attempting to open xlsx file: %s' % file_to_open)
            try:
                self.workbook = openpyxl.load_workbook(file_to_open,
                                                       data_only=True)
                self.filepath = file_to_open
            except openpyxl.exceptions.InvalidFileException as error:
                log.error(error)

    def skip_sheet(self, sheet_name):
        """Performs a case-insensitive search of *sheet_name*
        against :attr:`parser.skip_sheets`

        **Returns:**
            Boolean ``True`` if current sheet should be skipped.
            Boolean ``False`` otherwise

        """
        status = False

        if sheet_name.lower() in [x.lower() for x in self.skip_sheets]:
            log.debug('Sheet "%s" set to be skipped' % sheet_name)
            status = True

        return status

    def parse_sheets(self):
        """Will attempt to extract the cells defined by
        :attr:`cells_to_extract` from the worksheets within
        :attr:`workbook`

        **Returns:**
            dictionary structure representing the values parsed from the
            Excel worksheets.  Typical structure is::

                {<worksheet_name>: {<cell_to_extract>: <cell_value>}, ...}

            For example::

                {'AAA-000-001': {'B1': u'AAA-000-001'},
                 'CLM-121-001': {'B1': u'CLM-121-001'},
                 'Instructions': {'B1': None},
                 'WorkbookLog': {'B1': u'Date'}}

        """
        parsed_values = {}

        for sheet in self.workbook.get_sheet_names():
            log.info('Extracting from sheet name: "%s"' % sheet)
            if self.skip_sheet(sheet):
                continue

            # Need to make the key unique as the concatenation of the
            # workbook and worksheet
            key = '%s|%s' % (os.path.basename(self.filepath), sheet)
            parsed_values[key] = {}

            # Set active sheet.
            ws = self.workbook.get_sheet_by_name(sheet)

            # Extract required cells.
            for cell in self.cells_to_extract:
                value = ws[cell].value
                log.debug('Extracted cell|value: %s|%s' % (cell, value))
                parsed_values[key][cell] = value

        return parsed_values
